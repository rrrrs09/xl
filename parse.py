import sys
import datetime

import xlrd
from openpyxl import Workbook
from openpyxl.styles import (
    NamedStyle,
    Font, 
    Color,
    Alignment,
    Border,
    Side,
    PatternFill,
    fills,
    colors
)


def remove_blank_cells(row):
    return list(filter(lambda cell: cell != '', row))


def process_input_file(input_file_path):
    '''
    Получает на вход путь к входному файлу
    Возвращает список из словарей с сотрудником и расписанием
    '''
    try:
        rb = xlrd.open_workbook(input_file_path)
    except FileNotFoundError as error:
        print(f'Файл "{error.filename}" не найден')
        return
    sheet = rb.sheet_by_index(0)
    
    employees = []
    row_num = 0
    while row_num < sheet.nrows:
        # удаляем все пустые ячейки из строки
        row = remove_blank_cells(sheet.row_values(row_num))
        if not row:
            row_num += 1
            continue
        employee = {}
        if row[0] == 'Сотрудник':
            employee['name'] = row[1] if len(row) > 1 else '' 
            row = remove_blank_cells(sheet.row_values(row_num+2))
            if row[0] == 'Отдел':
                employee['department'] = row[1] if len(row) > 1 else ''
            row = remove_blank_cells(sheet.row_values(row_num+3))
            if row[0] == 'Должность':
                employee['position'] = row[1] if len(row) > 1 else ''
            row_num += 7
            employee['schedule'] = []
            row = remove_blank_cells(sheet.row_values(row_num))
            if row[0] == 'Дата':
                row_num += 1
                while True:
                    row = remove_blank_cells(sheet.row_values(row_num))
                    if not row:
                        row_num += 1
                        continue
                    if row[0] == 'Итого':
                        employee['total'] = row
                        row_num += 1
                        break
                    # Добавляем всю строку расписания
                    employee['schedule'].append(row)
                    row_num += 1
                
        row_num += 1
        if employee and employee['schedule']:
            employees.append(employee)
    return employees


def get_regular_number_of_days(schedule):
    '''
    Возвращает штатное количество дней
    Рассчитывается по столбцу "Норма"
    '''
    return len(list(filter(lambda day: day[4].strip() != '-', schedule)))


def get_actual_number_of_days(schedule):
    '''
    Возвращает фактическое количество дней по СКУД
    Рассчитывается по столбцу "Отбработка"
    '''
    return len(list(filter(lambda day: day[3].strip() != '-', schedule)))


def get_delta_between_indicators(actual_time, total_time):
    '''
    Принимает фактическое количество часов и количество часов отработки
    Возвращает Дельта между СКУД и штатными показателями
    '''
    actual_number_of_hours = datetime.timedelta(
        hours=int(actual_time.split(':')[0]), 
        minutes=int(actual_time.split(':')[1])
    )
    total_hours_count = datetime.timedelta(
        hours=int(total_time.split(':')[0]), 
        minutes=int(total_time.split(':')[1])
    )
    
    tot_sec = None
    delta_between_indicators = ''
    if actual_number_of_hours > total_hours_count:
        tot_sec = (actual_number_of_hours - total_hours_count).total_seconds()
        delta_between_indicators = '(переработка)'
    elif actual_number_of_hours < total_hours_count:
        tot_sec = (total_hours_count - actual_number_of_hours).total_seconds()
        delta_between_indicators = '(недоработка)'
    
    if tot_sec:
        hours = round(tot_sec // 3600)
        minutes = round((tot_sec % 3600) // 60)
        delta_between_indicators = f'{hours}:{minutes}:00 ' + delta_between_indicators
    
    return delta_between_indicators


def get_latenesses(schedule):
    '''Возвращет список дней, в которые сотрудник опоздал'''
    days = list(filter(lambda day: day[7].strip() != '-', schedule))
    return [f'{day[0]} ({day[7]}) - опоздание' for day in days]


def get_remark(latenesses):
    '''Возваращает примечание'''
    length = len(latenesses)
    if length == 0:
        return ''
    elif length % 10 == 1 and length != 11:
        return f'Штраф за {length} опоздание'
    elif 2 <= length <= 4 or 22 <= length <= 24:
        return f'Штраф за {length} опоздания'
    else:
        return f'Штраф за {length} опозданий'


def write_cell(ws, text, style , from_cell, to_cell=''):
    if to_cell != '':
        ws.merge_cells(f'{from_cell}:{to_cell}')
    ws[from_cell] = text
    ws[from_cell].style = style


def get_date_interval(schedule):
    '''
    Принимает расписание сотрудника
    Возвращает интервал в виде строки
    '''
    months = {
        '1': 'января',
        '2': 'февраля',
        '3': 'марта',
        '4': 'апреля',
        '5': 'мая',
        '6': 'июня',
        '7': 'июля',
        '8': 'августа',
        '9': 'сентября',
        '10': 'октября',
        '11': 'ноября',
        '12': 'декабря',
    }
    month = schedule[0][0].split('.')[1].lstrip('0')
    first_day = schedule[0][0].split('.')[0].lstrip('0')
    last_day = schedule[-1][0].split('.')[0]
    return f'с {first_day} по {last_day} {months[month]}'


def write_file(employees):
    wb = Workbook()
    ws = wb.active
    date_interval = get_date_interval(employees[0]['schedule'])
    
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 22
    ws.column_dimensions['I'].width = 37
    ws.column_dimensions['J'].width = 25
    
    font_style = Font(name='Times New Roman', size=12)
    bold_font_style = Font(name='Times New Roman', size=12, bold=True)
    right_align = Alignment(horizontal='right', vertical='center', wrapText=True)
    center_align = Alignment(horizontal='center', vertical='center', wrapText=True)
    left_align = Alignment(horizontal='left', vertical='center', wrapText=True)
    border_color = Color()
    border_side_medium = Side(color=border_color, border_style='medium')
    border_side_thin = Side(color=border_color, border_style='thin')
    
    table_head_border = Border(
        left=border_side_medium,
        right=border_side_medium,
        top=border_side_medium,
        bottom=border_side_medium
    )
    table_body_border = Border(
        left=border_side_thin,
        right=border_side_thin,
        top=border_side_thin,
        bottom=border_side_thin
    )
    table_head_fill = PatternFill(
        fgColor=colors.COLOR_INDEX[22],
        bgColor=colors.COLOR_INDEX[22], 
        fill_type=fills.FILL_SOLID
    )
    table_head_style = NamedStyle(
        'table head style', 
        font=bold_font_style, 
        alignment=center_align,
        border=table_head_border,
        fill=table_head_fill
    )
    table_body_style_left = NamedStyle(
        'table body style left', 
        font=font_style, 
        alignment=left_align,
        border=table_body_border
    )
    table_body_style_center = NamedStyle(
        'table body style center', 
        font=font_style, 
        alignment=center_align,
        border=table_body_border
    )
    
    write_cell(ws, from_cell='A4', to_cell='J4', text='Приложение №2', 
        style=NamedStyle('A4 style', font=bold_font_style, alignment=right_align)
    )
    write_cell(ws, from_cell='A5', to_cell='J5', text='К Приказу № 22 от 30.03.2018 г.', 
        style=NamedStyle('A5 style', font=bold_font_style, alignment=right_align)
    )
    write_cell(ws, from_cell='A6', to_cell='J6', text='«О изменении режима рабочего времени и', 
        style=NamedStyle('A6 style', font=font_style, alignment=right_align)
    )
    write_cell(ws, from_cell='A7', to_cell='J7', text='применении взысканий за нарушение', 
        style=NamedStyle('A7 style', font=font_style, alignment=right_align)
    )
    write_cell(ws, from_cell='A8', to_cell='J8', text='трудовой дисциплины ООО «Прогресс»', 
        style=NamedStyle('A8 style', font=font_style, alignment=right_align)
    )
    write_cell(ws, from_cell='A9', to_cell='J9',
        text='Табель учета рабочего времени ' + date_interval, 
        style=NamedStyle('A9 style', font=bold_font_style, alignment=center_align)
    )
    
    write_cell(ws, from_cell='A11', to_cell='A12', text='№ п/п', 
        style=table_head_style
    )
    ws['A12'].border = table_head_border
    write_cell(ws, from_cell='B11', to_cell='B12', text='Подразделение', 
        style=table_head_style
    )
    ws['B12'].border = table_head_border
    write_cell(ws, from_cell='C11', to_cell='C12', text='ФИО', 
        style=table_head_style
    )
    ws['C12'].border = table_head_border
    write_cell(ws, from_cell='D11', to_cell='D12', text='Должность', 
        style=table_head_style
    )
    ws['D12'].border = table_head_border
    write_cell(ws, from_cell='E11', text='Штатное количество', style=table_head_style)
    write_cell(ws, from_cell='E12', text='дней', style=table_head_style)
    write_cell(ws, from_cell='F11', to_cell='G11', text='Фактическое количество по СКУД', 
        style=table_head_style
    )
    ws['G11'].border = table_head_border
    write_cell(ws, from_cell='F12', text='дней', style=table_head_style)
    write_cell(ws, from_cell='G12', text='часов', style=table_head_style)
    write_cell(ws, from_cell='H11', 
        text='Дельта между СКУД и штатными показателями', 
        style=table_head_style
    )
    write_cell(ws, from_cell='H12', text='дней', style=table_head_style)
    write_cell(ws, from_cell='I11', to_cell='I12', text='Примечание', 
        style=table_head_style
    )
    ws['I12'].border = table_head_border
    write_cell(ws, from_cell='J11', to_cell='J12', 
        text='Заключение руководителя по выплате', 
        style=table_head_style
    )
    ws['J12'].border = table_head_border
    
    start_row = 13
    for (index, emp) in enumerate(employees):
        actual_number_of_days = get_actual_number_of_days(emp['schedule'])
        actual_number_of_hours = emp['total'][1].replace(':', ',')
        regular_number_of_days = get_regular_number_of_days(emp['schedule'])
        delta_between_indicators = get_delta_between_indicators(
            emp['total'][1],
            emp['total'][2]
        )
        latenesses = get_latenesses(emp['schedule'])
        remark = get_remark(latenesses)
        
        row_num = str(start_row+index)
        ws['A'+row_num] = index + 1
        ws['A'+row_num].style = table_body_style_center
        ws['B'+row_num] = emp['department']
        ws['B'+row_num].style = table_body_style_left
        ws['C'+row_num] = emp['name']
        ws['C'+row_num].style = table_body_style_left
        ws['D'+row_num] = emp['position']
        ws['D'+row_num].style = table_body_style_left
        ws['E'+row_num] = regular_number_of_days
        ws['E'+row_num].style = table_body_style_center
        ws['F'+row_num] = actual_number_of_days
        ws['F'+row_num].style = table_body_style_center
        ws['G'+row_num] = actual_number_of_hours
        ws['G'+row_num].style = table_body_style_center
        ws['H'+row_num] = delta_between_indicators
        ws['H'+row_num].style = table_body_style_center
        ws['I'+row_num] = '\n'.join(latenesses)
        ws['I'+row_num].style = table_body_style_left
        ws['J'+row_num] = remark
        ws['J'+row_num].style = table_body_style_center
        
    wb.save(f'Табель учета рабочего времени {date_interval}.xlsx')


if __name__ == '__main__':
    try:
        input_file_path = sys.argv[1]
    except IndexError as error:
        print('Путь к файлу не передан')
        exit()
    employees = process_input_file(input_file_path)
    if employees:
        write_file(employees)
